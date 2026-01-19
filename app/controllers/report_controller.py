import io
from flask import Blueprint, send_file
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from app.models.Models import get_monthly_and_yearly_revenue
from app.controllers.page_routes import roles_required

report_bp = Blueprint("report", __name__)

# Export báo cáo
@report_bp.route("/export-revenue", methods=["GET"])
@roles_required("Admin", "Accountant")
def export_revenue_excel():
    monthly_rows, year_total_map = get_monthly_and_yearly_revenue()
    wb = Workbook()
    ws = wb.active
    ws.title = "Doanh thu"

    headers = ["Tháng", "Năm", "Tổng doanh thu tháng", "Tổng doanh thu 12 tháng"]
    ws.append(headers)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, 5):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # Data
    for year, month, total in monthly_rows:
        year_total = float(year_total_map.get(year, 0))
        ws.append([month, year, float(total or 0), year_total])
    for r in range(2, ws.max_row + 1):
        ws[f"C{r}"].number_format = "#,##0"
        ws[f"D{r}"].number_format = "#,##0"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 24

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="bao_cao_doanh_thu_thang_nam.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
